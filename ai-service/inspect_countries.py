"""
Inspect the actual country data in the Excel file.
"""
from openpyxl import load_workbook


def inspect_countries():
    """Inspect country values in the Excel file."""
    input_file = r"c:\Users\FirdavsMuzaffarov\Desktop\agent-projects\compliance-ai-service\customers_example_list.xlsx"
    
    print(f"Loading: {input_file}\n")
    wb = load_workbook(input_file)
    sheet = wb.active
    
    # Find header row
    header_row = None
    for row_idx in range(1, 6):
        for cell in sheet[row_idx]:
            if cell.value and "customerno" in str(cell.value).lower().replace(" ", ""):
                header_row = row_idx
                break
        if header_row:
            break
    
    # Get column mapping
    columns = {}
    for cell in sheet[header_row]:
        if cell.value:
            columns[str(cell.value).strip()] = cell.column
    
    print("Country-related columns:")
    for col_name in ["BirthCountry", "Citizenship", "CitizenshipDesc", "Nationality", "NationalityDesc"]:
        if col_name in columns:
            print(f"  ✅ {col_name}")
    print()
    
    # Display country values for first 10 customers
    print("=" * 100)
    print(f"{'CustomerNo':<20} {'CitizenshipDesc':<25} {'Citizenship':<15} {'BirthCountry':<25}")
    print("=" * 100)
    
    customer_col = columns.get("CustomerNo")
    citizenship_desc_col = columns.get("CitizenshipDesc")
    citizenship_col = columns.get("Citizenship")
    birth_country_col = columns.get("BirthCountry")
    
    rows_shown = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        if rows_shown >= 10:
            break
        
        customer_no = sheet.cell(row_idx, customer_col).value if customer_col else ""
        if not customer_no:
            continue
        
        citizenship_desc = sheet.cell(row_idx, citizenship_desc_col).value if citizenship_desc_col else ""
        citizenship = sheet.cell(row_idx, citizenship_col).value if citizenship_col else ""
        birth_country = sheet.cell(row_idx, birth_country_col).value if birth_country_col else ""
        
        print(f"{str(customer_no):<20} {str(citizenship_desc):<25} {str(citizenship):<15} {str(birth_country):<25}")
        rows_shown += 1
    
    print("=" * 100)


if __name__ == "__main__":
    inspect_countries()
