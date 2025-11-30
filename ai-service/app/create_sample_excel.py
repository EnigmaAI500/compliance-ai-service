"""
Script to generate a sample Excel file for batch risk assessment testing.
Run this to create customers_example_list.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

def create_sample_excel():
    """Create a sample Excel file with test customer data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define headers - matching the actual database structure
    headers = [
        "CustomerNo",
        "DocumentName",
        "MainAccount",
        "District",
        "Region",
        "Locality",
        "Street",
        "Pinfl",
        "ExpiryDate",
        "Nationality",
        "BirthCountry",
        "PassportIssuerCode",
        "PassportIssuerPlace",
        "Citizenship",
        "RegDocType",
        "RegDocNum",
        "RegDocSerialNum",
        "RegPinfl",
        "Lang",
        "CitizenshipDesc",
        "NationalityDesc",
        "AddressCode",
        "ResidentStatus",
        "RiskFlag",
        "LocalBlackListFlag"
    ]
    
    # Write headers with styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Sample data - mix of risk levels
    sample_customers = [
        # High risk - FATF black list country
        {
            "CustomerNo": "CUST-001",
            "DocumentName": "Ahmad Rezaei",
            "MainAccount": "currency_exchange",
            "District": "Tehran",
            "Region": "Central",
            "Locality": "Downtown",
            "Street": "Main Street 123",
            "Pinfl": "12345678901234",
            "ExpiryDate": "2026-12-31",
            "Nationality": "Iranian",
            "BirthCountry": "Iran",
            "PassportIssuerCode": "IR",
            "PassportIssuerPlace": "Tehran",
            "Citizenship": "Iran",
            "RegDocType": "Passport",
            "RegDocNum": "A1234567",
            "RegDocSerialNum": "IR001",
            "RegPinfl": "12345678901234",
            "Lang": "fa",
            "CitizenshipDesc": "Iranian",
            "NationalityDesc": "Iranian",
            "AddressCode": "100001",
            "ResidentStatus": "Resident",
            "RiskFlag": "",
            "LocalBlackListFlag": "N"
        },
        # High risk - PEP + FATF grey list
        {
            "CustomerNo": "CUST-002",
            "DocumentName": "Maria Rodriguez",
            "MainAccount": "government_official",
            "District": "Caracas",
            "Region": "Capital",
            "Locality": "City Center",
            "Street": "Bolivar Ave 456",
            "Pinfl": "23456789012345",
            "ExpiryDate": "2027-06-30",
            "Nationality": "Venezuelan",
            "BirthCountry": "Venezuela",
            "PassportIssuerCode": "VE",
            "PassportIssuerPlace": "Caracas",
            "Citizenship": "Venezuela",
            "RegDocType": "Passport",
            "RegDocNum": "V7654321",
            "RegDocSerialNum": "VE002",
            "RegPinfl": "23456789012345",
            "Lang": "es",
            "CitizenshipDesc": "Venezuelan",
            "NationalityDesc": "Venezuelan",
            "AddressCode": "200002",
            "ResidentStatus": "Resident",
            "RiskFlag": "PEP",
            "LocalBlackListFlag": "N"
        },
        # Medium risk - FATF grey list
        {
            "CustomerNo": "CUST-003",
            "DocumentName": "John Smith",
            "MainAccount": "retail",
            "District": "Damascus",
            "Region": "South",
            "Locality": "Suburbs",
            "Street": "Market Road 789",
            "Pinfl": "34567890123456",
            "ExpiryDate": "2025-03-15",
            "Nationality": "Syrian",
            "BirthCountry": "Syria",
            "PassportIssuerCode": "SY",
            "PassportIssuerPlace": "Damascus",
            "Citizenship": "Syria",
            "RegDocType": "Passport",
            "RegDocNum": "S9876543",
            "RegDocSerialNum": "SY003",
            "RegPinfl": "34567890123456",
            "Lang": "ar",
            "CitizenshipDesc": "Syrian",
            "NationalityDesc": "Syrian",
            "AddressCode": "300003",
            "ResidentStatus": "Non-Resident",
            "RiskFlag": "",
            "LocalBlackListFlag": "N"
        },
        # Low risk - clean country
        {
            "CustomerNo": "CUST-004",
            "DocumentName": "Emma Johnson",
            "MainAccount": "software_engineer",
            "District": "London",
            "Region": "Southeast",
            "Locality": "Westminster",
            "Street": "Oxford Street 321",
            "Pinfl": "45678901234567",
            "ExpiryDate": "2028-09-20",
            "Nationality": "British",
            "BirthCountry": "United Kingdom",
            "PassportIssuerCode": "GB",
            "PassportIssuerPlace": "London",
            "Citizenship": "United Kingdom",
            "RegDocType": "Passport",
            "RegDocNum": "GB123456",
            "RegDocSerialNum": "GB004",
            "RegPinfl": "45678901234567",
            "Lang": "en",
            "CitizenshipDesc": "British",
            "NationalityDesc": "British",
            "AddressCode": "400004",
            "ResidentStatus": "Resident",
            "RiskFlag": "",
            "LocalBlackListFlag": "N"
        },
        # High risk - Local blacklist
        {
            "CustomerNo": "CUST-005",
            "DocumentName": "Viktor Petrov",
            "MainAccount": "casino",
            "District": "Moscow",
            "Region": "Central",
            "Locality": "Downtown",
            "Street": "Red Square 555",
            "Pinfl": "56789012345678",
            "ExpiryDate": "2026-01-10",
            "Nationality": "Russian",
            "BirthCountry": "Russia",
            "PassportIssuerCode": "RU",
            "PassportIssuerPlace": "Moscow",
            "Citizenship": "Russia",
            "RegDocType": "Passport",
            "RegDocNum": "RU789012",
            "RegDocSerialNum": "RU005",
            "RegPinfl": "56789012345678",
            "Lang": "ru",
            "CitizenshipDesc": "Russian",
            "NationalityDesc": "Russian",
            "AddressCode": "500005",
            "ResidentStatus": "Resident",
            "RiskFlag": "",
            "LocalBlackListFlag": "Y"
        }
    ]
    
    # Write data rows
    for row_idx, customer in enumerate(sample_customers, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = customer.get(header, "")
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    
    # Save file
    output_path = Path(__file__).parent.parent / "customers_example_list.xlsx"
    wb.save(output_path)
    print(f"✅ Sample Excel file created: {output_path}")
    print(f"   Contains {len(sample_customers)} sample customers")
    print(f"   Ready for upload to /risk/batch-excel endpoint")
    
    return output_path


if __name__ == "__main__":
    create_sample_excel()
