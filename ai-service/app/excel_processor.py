"""
Excel batch processing utilities for customer risk assessment.
Handles reading customer data from Excel, processing risk scores, and writing results back.
"""
import io
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .pure_risk_scorer import PureRiskScorer
from .logging_config import get_failure_logger


logger = get_failure_logger()


class ExcelCustomerProcessor:
    """
    Processes customer risk assessment from Excel files.
    Expected columns: CustomerNo, FirstName, LastName, BirthCountry, Citizenship, 
                     Nationality, MainAccount, RiskFlag (optional PEP flag)
    Adds columns: RiskScore, RiskFlag, RiskReason
    """
    
    # Required input columns
    REQUIRED_COLUMNS = [
        "CustomerNo",
        "DocumentName",
        "BirthCountry",
        "Citizenship"
    ]
    
    # Optional input columns
    OPTIONAL_COLUMNS = [
        "Nationality",
        "MainAccount",
        "RiskFlag",
        "LocalBlackListFlag",
        "ResidentStatus",
        "District",
        "Region",
        "Locality",
        "Street",
        "Pinfl",
        "PassportIssuerCode",
        "PassportIssuerPlace",
        "NationalityDesc",
        "CitizenshipDesc"
    ]
    
    # Output columns we'll add
    OUTPUT_COLUMNS = [
        "RiskScore",
        "RiskFlag",
        "RiskReason"
    ]
    
    def __init__(self, use_llm: bool = False):
        """
        Initialize processor.
        
        Args:
            use_llm: Deprecated parameter, kept for compatibility. Pure algorithm is always used.
        """
        self.risk_scorer = PureRiskScorer()
    
    def _normalize_column_name(self, name: str) -> str:
        """Normalize column names (trim, title case)."""
        if not name:
            return ""
        return str(name).strip()
    
    def _find_header_row(self, sheet: Worksheet) -> Optional[int]:
        """
        Find the row containing column headers.
        Looks for 'CustomerNo' in the first 5 rows.
        
        Returns:
            Row number (1-indexed) or None if not found
        """
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            for cell in sheet[row_idx]:
                if cell.value and "customerno" in str(cell.value).lower().replace(" ", ""):
                    return row_idx
        return None
    
    def _get_column_mapping(self, sheet: Worksheet, header_row: int) -> Dict[str, int]:
        """
        Create mapping of column names to column indices.
        
        Returns:
            Dict mapping normalized column names to 1-indexed column numbers
        """
        mapping = {}
        for cell in sheet[header_row]:
            if cell.value:
                col_name = self._normalize_column_name(cell.value)
                mapping[col_name] = cell.column
        return mapping
    
    def _validate_columns(self, column_mapping: Dict[str, int]) -> List[str]:
        """
        Validate that all required columns are present.
        
        Returns:
            List of missing column names (empty if all present)
        """
        missing = []
        for required in self.REQUIRED_COLUMNS:
            if required not in column_mapping:
                missing.append(required)
        return missing
    
    def _extract_row_data(
        self, 
        sheet: Worksheet, 
        row_idx: int, 
        column_mapping: Dict[str, int]
    ) -> Dict[str, Any]:
        """
        Extract customer data from a single row.
        
        Args:
            sheet: Excel worksheet
            row_idx: Row index (1-indexed)
            column_mapping: Column name to column index mapping
            
        Returns:
            Dictionary with customer data
        """
        data = {}
        for col_name, col_idx in column_mapping.items():
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            data[col_name] = cell_value if cell_value is not None else ""
        return data
    
    def _row_to_profile(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Convert Excel row data to risk assessment profile format.
        
        Args:
            row_data: Raw data from Excel row
            
        Returns:
            Profile dict compatible with llm_score_sanctions()
        """
        return {
            "country_of_birth": str(row_data.get("BirthCountry") or ""),
            "residency_country": str(row_data.get("Citizenship") or row_data.get("Nationality") or ""),
            "occupation": str(row_data.get("MainAccount") or "").lower(),
            "is_pep": str(row_data.get("RiskFlag", "")).upper() == "PEP",
            "full_name": str(row_data.get("DocumentName") or "").strip(),
            
            # Additional fields
            "nationality": str(row_data.get("Nationality") or ""),
            "citizenship": str(row_data.get("Citizenship") or ""),
            "resident_status": str(row_data.get("ResidentStatus") or ""),
            "district": str(row_data.get("District") or ""),
            "region": str(row_data.get("Region") or ""),
            "locality": str(row_data.get("Locality") or ""),
        }
    
    def _assess_customer_risk(
        self, 
        profile: Dict[str, Any], 
        local_blacklist_flag: bool
    ) -> Dict[str, Any]:
        """
        Perform risk assessment for a single customer.
        
        Args:
            profile: Customer profile dict (original row_data, not transformed)
            local_blacklist_flag: Whether customer is on local blacklist
            
        Returns:
            Dict with risk_score, risk_band, reasons
        """
        try:
            # Add LocalBlackListFlag to profile if not already there
            if "LocalBlackListFlag" not in profile:
                profile["LocalBlackListFlag"] = "Y" if local_blacklist_flag else "N"
            
            # Use pure algorithm risk scoring
            result = self.risk_scorer.assess_risk(profile)
            
            risk_score = int(result.get("risk_score", 0))
            risk_flag = str(result.get("risk_flag", "GREEN")).upper()
            risk_reason = str(result.get("risk_reason", "No assessment performed"))
            
            # Convert single reason string to list for consistency
            reasons = [risk_reason]
            
            return {
                "risk_score": risk_score,
                "risk_band": risk_flag,
                "reasons": reasons
            }
            
        except Exception as ex:
            logger.exception("Error assessing risk for customer: %s", ex)
            return {
                "risk_score": 50,
                "risk_band": "YELLOW",
                "reasons": [f"ERROR_DURING_ASSESSMENT: {type(ex).__name__}"]
            }
    
    def _add_output_columns(
        self, 
        sheet: Worksheet, 
        header_row: int, 
        column_mapping: Dict[str, int]
    ) -> Dict[str, int]:
        """
        Add output columns (RiskScore, RiskFlag, RiskReason) to the sheet.
        Updates column_mapping in place.
        
        Returns:
            Updated column mapping
        """
        next_col = max(column_mapping.values()) + 1
        
        for output_col in self.OUTPUT_COLUMNS:
            if output_col not in column_mapping:
                sheet.cell(row=header_row, column=next_col, value=output_col)
                column_mapping[output_col] = next_col
                next_col += 1
        
        return column_mapping
    
    def _write_results(
        self,
        sheet: Worksheet,
        row_idx: int,
        column_mapping: Dict[str, int],
        results: Dict[str, Any]
    ):
        """
        Write risk assessment results to Excel row.
        
        Args:
            sheet: Excel worksheet
            row_idx: Row index to write to
            column_mapping: Column name to index mapping
            results: Risk assessment results
        """
        if "RiskScore" in column_mapping:
            sheet.cell(
                row=row_idx, 
                column=column_mapping["RiskScore"], 
                value=results["risk_score"]
            )
        
        if "RiskFlag" in column_mapping:
            sheet.cell(
                row=row_idx, 
                column=column_mapping["RiskFlag"], 
                value=results["risk_band"]
            )
        
        if "RiskReason" in column_mapping:
            reasons_text = "; ".join(results["reasons"])
            sheet.cell(
                row=row_idx, 
                column=column_mapping["RiskReason"], 
                value=reasons_text
            )
    
    def process_excel_file(self, file_bytes: bytes) -> bytes:
        """
        Main processing function: reads Excel, assesses all customers, writes results.
        
        Args:
            file_bytes: Raw bytes of uploaded Excel file
            
        Returns:
            Bytes of processed Excel file with risk assessment results
            
        Raises:
            ValueError: If file format is invalid or required columns missing
        """
        # Load workbook
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
        except Exception as ex:
            raise ValueError(f"Invalid Excel file format: {ex}")
        
        # Get active sheet
        sheet = wb.active
        
        if not sheet or sheet.max_row < 2:
            raise ValueError("Excel file is empty or has no data rows")
        
        # Find header row
        header_row = self._find_header_row(sheet)
        if not header_row:
            raise ValueError("Could not find header row with 'CustomerNo' column")
        
        # Get column mapping
        column_mapping = self._get_column_mapping(sheet, header_row)
        
        # Validate required columns
        missing_cols = self._validate_columns(column_mapping)
        if missing_cols:
            raise ValueError(f"Missing required columns: {', '.join(missing_cols)}")
        
        # Add output columns
        column_mapping = self._add_output_columns(sheet, header_row, column_mapping)
        
        # Process each data row
        processed_count = 0
        error_count = 0
        
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            try:
                # Check if row is empty (CustomerNo is empty)
                customer_no_col = column_mapping["CustomerNo"]
                customer_no = sheet.cell(row=row_idx, column=customer_no_col).value
                
                if not customer_no:
                    continue  # Skip empty rows
                
                # Extract row data
                row_data = self._extract_row_data(sheet, row_idx, column_mapping)
                
                # Check local blacklist flag
                local_blacklist = str(row_data.get("LocalBlackListFlag", "")).upper() == "Y"
                
                # Assess risk (pass row_data directly to scorer)
                results = self._assess_customer_risk(row_data, local_blacklist)
                
                # Write results
                self._write_results(sheet, row_idx, column_mapping, results)
                
                processed_count += 1
                
                # Log progress every 10 rows
                if processed_count % 10 == 0:
                    logger.info(f"Processed {processed_count} customers...")
                
            except Exception as ex:
                error_count += 1
                logger.exception(f"Error processing row {row_idx}: {ex}")
                
                # Write error to results
                self._write_results(
                    sheet, 
                    row_idx, 
                    column_mapping, 
                    {
                        "risk_score": 50,
                        "risk_band": "YELLOW",
                        "reasons": [f"PROCESSING_ERROR: {type(ex).__name__}"]
                    }
                )
        
        logger.info(f"Batch processing complete: {processed_count} customers processed, {error_count} errors")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()


def process_excel_batch(file_bytes: bytes, use_llm: bool = True) -> bytes:
    """
    Convenience function to process Excel file with risk assessments.
    
    Args:
        file_bytes: Raw bytes of Excel file
        use_llm: Whether to use LLM for assessments
        
    Returns:
        Processed Excel file as bytes
    """
    processor = ExcelCustomerProcessor(use_llm=use_llm)
    return processor.process_excel_file(file_bytes)
