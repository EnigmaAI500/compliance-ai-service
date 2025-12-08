import io
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .logging_config import get_failure_logger

logger = get_failure_logger()


class BulkExcelParser:
    # Updated input columns (your new template)
    REQUIRED_COLUMNS = [
        "CustomerNo",
        "DocumentName",
        "MainAccount",
        "District",
        "Region",
        "Locality",
        "Street",
        "Pinfl",
        "ExpiryDate",
        "Nationality",
        "BirthCountry",
        "PassportIssuerCode",
        "PassportIssuerPlace",
        "Citizenship",
        "RegDocType",
        "RegDocNum",
        "RegDocSerialNum",
        "RegPinfl",
        "Lang",
        "CitizenshipDesc",
        "NationalityDesc",
        "AddressCode",
        "ResidentStatus",
        # plus your 2 new columns:
        "Email",
        "IPAddress",
        # RiskFlag/RiskScore/RiskReason treated as optional, see below
    ]

    OPTIONAL_COLUMNS = [
        "RiskFlag",
        "RiskScore",
        "RiskReason",
    ]

    def _normalize_name(self, name: str) -> str:
        return str(name or "").strip()

    def _find_header_row(self, sheet: Worksheet) -> Optional[int]:
        # reuse logic from ExcelCustomerProcessor._find_header_row
        for row_idx in range(1, min(6, sheet.max_row + 1)):
            for cell in sheet[row_idx]:
                if cell.value and "customerno" in str(cell.value).lower().replace(" ", ""):
                    return row_idx
        return None

    def _get_column_mapping(self, sheet: Worksheet, header_row: int) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for cell in sheet[header_row]:
            if cell.value:
                col_name = self._normalize_name(cell.value)
                mapping[col_name] = cell.column
        return mapping

    def _extract_row_data(
        self, sheet: Worksheet, row_idx: int, column_mapping: Dict[str, int]
    ) -> Dict[str, Any]:
        data: Dict[str, Any] = {}
        for col_name, col_idx in column_mapping.items():
            value = sheet.cell(row=row_idx, column=col_idx).value
            data[col_name] = value if value is not None else ""
        return data

    def parse(self, file_bytes: bytes) -> Dict[str, Any]:
        # open workbook
        try:
            wb = load_workbook(io.BytesIO(file_bytes))
        except Exception as ex:
            raise ValueError(f"Invalid Excel file format: {ex}")

        sheet = wb.active
        if not sheet or sheet.max_row < 2:
            raise ValueError("Excel file is empty or has no data rows")

        header_row = self._find_header_row(sheet)
        if not header_row:
            raise ValueError("Could not find header row with 'CustomerNo'")

        column_mapping = self._get_column_mapping(sheet, header_row)

        # validation
        missing_required = [c for c in self.REQUIRED_COLUMNS if c not in column_mapping]
        if missing_required:
            raise ValueError(f"Missing required columns: {', '.join(missing_required)}")

        missing_optional = [c for c in self.OPTIONAL_COLUMNS if c not in column_mapping]

        customers: List[Dict[str, Any]] = []
        rows_processed = 0

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            customer_no_col = column_mapping["CustomerNo"]
            customer_no_val = sheet.cell(row=row_idx, column=customer_no_col).value
            if not customer_no_val:
                continue

            raw_row = self._extract_row_data(sheet, row_idx, column_mapping)
            rows_processed += 1

            # build risk_input object (shape from your spec)
            customers.append(self._row_to_risk_input(raw_row))

        return {
            "rows_processed": rows_processed,
            "missing_optional": missing_optional,
            "customers": customers,
        }

    def _row_to_risk_input(self, row: Dict[str, Any]) -> Dict[str, Any]:
        # this matches the JSON in your description
        return {
            "customerNo": str(row.get("CustomerNo") or "").strip(),
            "fullName": str(row.get("DocumentName") or "").strip(),
            "citizenship": str(row.get("Citizenship") or ""),
            "citizenshipDesc": str(row.get("CitizenshipDesc") or ""),
            "nationality": str(row.get("Nationality") or ""),
            "nationalityDesc": str(row.get("NationalityDesc") or ""),
            "birthCountry": str(row.get("BirthCountry") or ""),

            "document": {
                "type": str(row.get("RegDocType") or ""),
                "number": str(row.get("RegDocNum") or ""),
                "serial": str(row.get("RegDocSerialNum") or ""),
                "issuerCode": str(row.get("PassportIssuerCode") or ""),
                "issuerPlace": str(row.get("PassportIssuerPlace") or ""),
                "expiryDate": str(row.get("ExpiryDate") or ""),
            },

            "residency": {
                "residentStatus": str(row.get("ResidentStatus") or ""),
                "district": str(row.get("District") or ""),
                "region": str(row.get("Region") or ""),
                "locality": str(row.get("Locality") or ""),
                "street": str(row.get("Street") or ""),
                "addressCode": str(row.get("AddressCode") or ""),
            },

            "business": {
                "mainAccount": str(row.get("MainAccount") or "").lower()
            },

            "digital": {
                "email": str(row.get("Email") or ""),
                "ipAddress": str(row.get("IPAddress") or ""),
            },

            "sourceFlags": {
                "riskFlag": str(row.get("RiskFlag") or "") or None,
                "prevRiskScore": row.get("RiskScore"),
                "prevRiskReason": str(row.get("RiskReason") or "") or None,
            },

            "rawRow": row,  # keep full original row for rawInput
        }
