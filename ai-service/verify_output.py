"""
Verify the processed Excel file to see risk scores and flags.
"""
from openpyxl import load_workbook


def verify_output():
    """Verify the processed Excel file."""
    output_file = r"c:\Users\FirdavsMuzaffarov\Desktop\agent-projects\compliance-ai-service\customers_RESULT.xlsx"
    
    print(f"Loading: {output_file}\n")
    wb = load_workbook(output_file)
    sheet = wb.active
    
    # Find header row and columns
    header_row = None
    for row_idx in range(1, 6):
        for cell in sheet[row_idx]:
            if cell.value and "customerno" in str(cell.value).lower().replace(" ", ""):
                header_row = row_idx
                break
        if header_row:
            break
    
    if not header_row:
        print("❌ Could not find header row")
        return
    
    # Get column mapping
    columns = {}
    for cell in sheet[header_row]:
        if cell.value:
            columns[str(cell.value).strip()] = cell.column
    
    print("Columns found:", list(columns.keys()))
    print()
    
    # Check required output columns
    required_output = ["RiskScore", "RiskFlag", "RiskReason"]
    for col in required_output:
        if col not in columns:
            print(f"❌ Missing column: {col}")
        else:
            print(f"✅ Found column: {col}")
    print()
    
    # Display sample rows
    print("=" * 120)
    print(f"{'CustomerNo':<15} {'Name':<25} {'Country':<20} {'RiskScore':<12} {'RiskFlag':<10} {'RiskReason':<30}")
    print("=" * 120)
    
    customer_col = columns.get("CustomerNo")
    name_col = columns.get("DocumentName")
    country_col = columns.get("CitizenshipDesc") or columns.get("Citizenship")
    score_col = columns.get("RiskScore")
    flag_col = columns.get("RiskFlag")
    reason_col = columns.get("RiskReason")
    
    rows_shown = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        if rows_shown >= 10:  # Show first 10 data rows
            break
        
        customer_no = sheet.cell(row_idx, customer_col).value if customer_col else ""
        if not customer_no:
            continue
        
        name = sheet.cell(row_idx, name_col).value if name_col else ""
        country = sheet.cell(row_idx, country_col).value if country_col else ""
        score = sheet.cell(row_idx, score_col).value if score_col else ""
        flag = sheet.cell(row_idx, flag_col).value if flag_col else ""
        reason = sheet.cell(row_idx, reason_col).value if reason_col else ""
        
        # Truncate long reason
        if reason and len(str(reason)) > 30:
            reason = str(reason)[:27] + "..."
        
        print(f"{str(customer_no):<15} {str(name):<25} {str(country):<20} {str(score):<12} {str(flag):<10} {str(reason):<30}")
        rows_shown += 1
    
    print("=" * 120)
    print(f"\n✅ Verified {rows_shown} customers")
    
    # Statistics
    stats = {"GREEN": 0, "YELLOW": 0, "RED": 0}
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        flag = sheet.cell(row_idx, flag_col).value if flag_col else ""
        if flag in stats:
            stats[flag] += 1
    
    print(f"\nRisk Distribution:")
    print(f"  GREEN:  {stats['GREEN']}")
    print(f"  YELLOW: {stats['YELLOW']}")
    print(f"  RED:    {stats['RED']}")


if __name__ == "__main__":
    verify_output()
